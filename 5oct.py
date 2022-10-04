from openpyxl import Workbook
from pathlib import Path
import os

print("Hello")

path = str(Path.cwd()) + "/sysInfoData/"
dir_list = os.listdir(path)

compList = []
for file in dir_list:
    with open(path+file, 'r') as f:
        specsList = f.readlines()
        compList.append(specsList)

wb = Workbook()
ws = wb['Sheet']

for i in range(len(dir_list)):
    hostname = compList[i][1].split(": ")[1].strip("\n")
    ip = compList[i][0].split(": ")[1].strip("\n")
    ws.cell(i+1, 1).value = hostname
    ws.cell(i+1, 2).value = ip

wb.save("sysInfoSheet.xlsx")
